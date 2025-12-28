"""
Módulo de generación de reportes (Excel y PDF)
"""
import io
from datetime import datetime
from typing import List, Dict, Any, Optional
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch

from .db import get_db


def generate_contracts_excel(filters: Optional[Dict] = None) -> bytes:
    """Genera reporte Excel de contratos"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Contratos"
    
    # Encabezados
    headers = ["ID", "Fecha", "Número Contrato", "Número ID", "Tipo", "Cliente", "RUC", "Estado"]
    ws.append(headers)
    
    # Estilo encabezados
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    
    # Consulta
    conn = get_db()
    cursor = conn.cursor()
    
    query = """
        SELECT c.id, c.fecha, c.numero_contrato, c.numero_id, c.tipo_contrato,
               cl.nombre as cliente_nombre, cl.ruc, c.estado
        FROM contracts c
        LEFT JOIN clients cl ON c.client_id = cl.id
        WHERE 1=1
    """
    params = []
    
    if filters:
        if filters.get('cliente'):
            query += " AND cl.nombre LIKE ?"
            params.append(f"%{filters['cliente']}%")
        if filters.get('numero_contrato'):
            query += " AND c.numero_contrato LIKE ?"
            params.append(f"%{filters['numero_contrato']}%")
        if filters.get('numero_id'):
            query += " AND c.numero_id LIKE ?"
            params.append(f"%{filters['numero_id']}%")
        if filters.get('estado'):
            query += " AND c.estado = ?"
            params.append(filters['estado'])
    
    query += " ORDER BY c.fecha DESC, c.id DESC"
    
    cursor.execute(query, params)
    rows = cursor.fetchall()
    conn.close()
    
    # Agregar datos
    for row in rows:
        ws.append([
            row['id'],
            row['fecha'],
            row['numero_contrato'],
            row['numero_id'] or '',
            row['tipo_contrato'] or '',
            row['cliente_nombre'] or '',
            row['ruc'] or '',
            row['estado']
        ])
    
    # Ajustar ancho de columnas
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Guardar en bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


def generate_contracts_pdf(filters: Optional[Dict] = None) -> bytes:
    """Genera reporte PDF de contratos"""
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    elements = []
    
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=16,
        textColor=colors.HexColor('#2c3e50'),
        spaceAfter=30,
    )
    
    # Título
    elements.append(Paragraph("Reporte de Contratos", title_style))
    elements.append(Spacer(1, 0.2*inch))
    
    # Consulta
    conn = get_db()
    cursor = conn.cursor()
    
    query = """
        SELECT c.id, c.fecha, c.numero_contrato, c.numero_id, c.tipo_contrato,
               cl.nombre as cliente_nombre, cl.ruc, c.estado
        FROM contracts c
        LEFT JOIN clients cl ON c.client_id = cl.id
        WHERE 1=1
    """
    params = []
    
    if filters:
        if filters.get('cliente'):
            query += " AND cl.nombre LIKE ?"
            params.append(f"%{filters['cliente']}%")
        if filters.get('numero_contrato'):
            query += " AND c.numero_contrato LIKE ?"
            params.append(f"%{filters['numero_contrato']}%")
        if filters.get('numero_id'):
            query += " AND c.numero_id LIKE ?"
            params.append(f"%{filters['numero_id']}%")
        if filters.get('estado'):
            query += " AND c.estado = ?"
            params.append(filters['estado'])
    
    query += " ORDER BY c.fecha DESC, c.id DESC"
    
    cursor.execute(query, params)
    rows = cursor.fetchall()
    conn.close()
    
    # Tabla
    data = [["ID", "Fecha", "Número", "Cliente", "RUC", "Estado"]]
    
    for row in rows:
        data.append([
            str(row['id']),
            row['fecha'],
            row['numero_contrato'],
            row['cliente_nombre'] or '',
            row['ruc'] or '',
            row['estado']
        ])
    
    table = Table(data)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    
    elements.append(table)
    doc.build(elements)
    buffer.seek(0)
    return buffer.getvalue()


def generate_purchase_orders_excel(filters: Optional[Dict] = None) -> bytes:
    """Genera reporte Excel de órdenes de compra"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Órdenes de Compra"
    
    headers = ["ID", "Fecha", "Número", "Contrato", "Cliente", "RUC", "Modo"]
    ws.append(headers)
    
    # Estilo encabezados
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    
    conn = get_db()
    cursor = conn.cursor()
    
    query = """
        SELECT po.id, po.fecha, po.numero, c.numero_contrato, cl.nombre as cliente_nombre,
               cl.ruc, po.sync_mode
        FROM purchase_orders po
        LEFT JOIN contracts c ON po.contract_id = c.id
        LEFT JOIN clients cl ON po.client_id = cl.id OR c.client_id = cl.id
        WHERE 1=1
    """
    params = []
    
    if filters:
        if filters.get('cliente'):
            query += " AND cl.nombre LIKE ?"
            params.append(f"%{filters['cliente']}%")
        if filters.get('contract_id'):
            query += " AND po.contract_id = ?"
            params.append(filters['contract_id'])
        if filters.get('id'):
            query += " AND po.id = ?"
            params.append(filters['id'])
    
    query += " ORDER BY po.fecha DESC, po.id DESC"
    
    cursor.execute(query, params)
    rows = cursor.fetchall()
    conn.close()
    
    for row in rows:
        ws.append([
            row['id'],
            row['fecha'],
            row['numero'],
            row['numero_contrato'] or '',
            row['cliente_nombre'] or '',
            row['ruc'] or '',
            row['sync_mode']
        ])
    
    # Ajustar columnas
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


def generate_purchase_orders_pdf(filters: Optional[Dict] = None) -> bytes:
    """Genera reporte PDF de órdenes de compra"""
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    elements = []
    
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=16,
        textColor=colors.HexColor('#2c3e50'),
        spaceAfter=30,
    )
    
    elements.append(Paragraph("Reporte de Órdenes de Compra", title_style))
    elements.append(Spacer(1, 0.2*inch))
    
    conn = get_db()
    cursor = conn.cursor()
    
    query = """
        SELECT po.id, po.fecha, po.numero, c.numero_contrato, cl.nombre as cliente_nombre,
               cl.ruc, po.sync_mode
        FROM purchase_orders po
        LEFT JOIN contracts c ON po.contract_id = c.id
        LEFT JOIN clients cl ON po.client_id = cl.id OR c.client_id = cl.id
        WHERE 1=1
    """
    params = []
    
    if filters:
        if filters.get('cliente'):
            query += " AND cl.nombre LIKE ?"
            params.append(f"%{filters['cliente']}%")
        if filters.get('contract_id'):
            query += " AND po.contract_id = ?"
            params.append(filters['contract_id'])
        if filters.get('id'):
            query += " AND po.id = ?"
            params.append(filters['id'])
    
    query += " ORDER BY po.fecha DESC, po.id DESC"
    
    cursor.execute(query, params)
    rows = cursor.fetchall()
    conn.close()
    
    data = [["ID", "Fecha", "Número", "Contrato", "Cliente", "RUC"]]
    
    for row in rows:
        data.append([
            str(row['id']),
            row['fecha'],
            row['numero'],
            row['numero_contrato'] or '',
            row['cliente_nombre'] or '',
            row['ruc'] or ''
        ])
    
    table = Table(data)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    
    elements.append(table)
    doc.build(elements)
    buffer.seek(0)
    return buffer.getvalue()


def generate_delivery_notes_excel(filters: Optional[Dict] = None) -> bytes:
    """Genera reporte Excel de notas de entrega"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Notas de Entrega"
    
    headers = ["ID", "Fecha", "Número", "Contrato", "Cliente", "RUC", "Dirección Entrega"]
    ws.append(headers)
    
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    
    conn = get_db()
    cursor = conn.cursor()
    
    query = """
        SELECT dn.id, dn.fecha, dn.numero_nota, c.numero_contrato, cl.nombre as cliente_nombre,
               cl.ruc, dn.direccion_entrega
        FROM delivery_notes dn
        LEFT JOIN contracts c ON dn.contract_id = c.id
        LEFT JOIN clients cl ON dn.client_id = cl.id OR c.client_id = cl.id
        WHERE 1=1
    """
    params = []
    
    if filters:
        if filters.get('cliente'):
            query += " AND cl.nombre LIKE ?"
            params.append(f"%{filters['cliente']}%")
        if filters.get('contract_id'):
            query += " AND dn.contract_id = ?"
            params.append(filters['contract_id'])
    
    query += " ORDER BY dn.fecha DESC, dn.id DESC"
    
    cursor.execute(query, params)
    rows = cursor.fetchall()
    conn.close()
    
    for row in rows:
        ws.append([
            row['id'],
            row['fecha'],
            row['numero_nota'],
            row['numero_contrato'] or '',
            row['cliente_nombre'] or '',
            row['ruc'] or '',
            row['direccion_entrega'] or ''
        ])
    
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


def generate_delivery_notes_pdf(filters: Optional[Dict] = None) -> bytes:
    """Genera reporte PDF de notas de entrega"""
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    elements = []
    
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=16,
        textColor=colors.HexColor('#2c3e50'),
        spaceAfter=30,
    )
    
    elements.append(Paragraph("Reporte de Notas de Entrega", title_style))
    elements.append(Spacer(1, 0.2*inch))
    
    conn = get_db()
    cursor = conn.cursor()
    
    query = """
        SELECT dn.id, dn.fecha, dn.numero_nota, c.numero_contrato, cl.nombre as cliente_nombre,
               cl.ruc, dn.direccion_entrega
        FROM delivery_notes dn
        LEFT JOIN contracts c ON dn.contract_id = c.id
        LEFT JOIN clients cl ON dn.client_id = cl.id OR c.client_id = cl.id
        WHERE 1=1
    """
    params = []
    
    if filters:
        if filters.get('cliente'):
            query += " AND cl.nombre LIKE ?"
            params.append(f"%{filters['cliente']}%")
        if filters.get('contract_id'):
            query += " AND dn.contract_id = ?"
            params.append(filters['contract_id'])
    
    query += " ORDER BY dn.fecha DESC, dn.id DESC"
    
    cursor.execute(query, params)
    rows = cursor.fetchall()
    conn.close()
    
    data = [["ID", "Fecha", "Número", "Contrato", "Cliente", "RUC"]]
    
    for row in rows:
        data.append([
            str(row['id']),
            row['fecha'],
            str(row['numero_nota']),
            row['numero_contrato'] or '',
            row['cliente_nombre'] or '',
            row['ruc'] or ''
        ])
    
    table = Table(data)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    
    elements.append(table)
    doc.build(elements)
    buffer.seek(0)
    return buffer.getvalue()


def generate_remissions_excel(filters: Optional[Dict] = None) -> bytes:
    """Genera reporte Excel de remisiones"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Remisiones"
    
    headers = ["ID", "Número", "Fecha Inicio", "Partida", "Llegada", "Transportista", "Cliente"]
    ws.append(headers)
    
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    
    conn = get_db()
    cursor = conn.cursor()
    
    query = """
        SELECT r.id, r.numero_remision, r.fecha_inicio, r.partida, r.llegada,
               r.transportista_nombre, cl.nombre as cliente_nombre
        FROM remissions r
        LEFT JOIN contracts c ON r.contract_id = c.id
        LEFT JOIN clients cl ON r.client_id = cl.id OR c.client_id = cl.id
        WHERE 1=1
    """
    params = []
    
    if filters:
        if filters.get('cliente'):
            query += " AND cl.nombre LIKE ?"
            params.append(f"%{filters['cliente']}%")
        if filters.get('contract_id'):
            query += " AND r.contract_id = ?"
            params.append(filters['contract_id'])
    
    query += " ORDER BY r.fecha_inicio DESC, r.id DESC"
    
    cursor.execute(query, params)
    rows = cursor.fetchall()
    conn.close()
    
    for row in rows:
        ws.append([
            row['id'],
            row['numero_remision'],
            row['fecha_inicio'],
            row['partida'] or '',
            row['llegada'] or '',
            row['transportista_nombre'] or '',
            row['cliente_nombre'] or ''
        ])
    
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


def generate_remissions_pdf(filters: Optional[Dict] = None) -> bytes:
    """Genera reporte PDF de remisiones"""
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    elements = []
    
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=16,
        textColor=colors.HexColor('#2c3e50'),
        spaceAfter=30,
    )
    
    elements.append(Paragraph("Reporte de Remisiones", title_style))
    elements.append(Spacer(1, 0.2*inch))
    
    conn = get_db()
    cursor = conn.cursor()
    
    query = """
        SELECT r.id, r.numero_remision, r.fecha_inicio, r.partida, r.llegada,
               r.transportista_nombre, cl.nombre as cliente_nombre
        FROM remissions r
        LEFT JOIN contracts c ON r.contract_id = c.id
        LEFT JOIN clients cl ON r.client_id = cl.id OR c.client_id = cl.id
        WHERE 1=1
    """
    params = []
    
    if filters:
        if filters.get('cliente'):
            query += " AND cl.nombre LIKE ?"
            params.append(f"%{filters['cliente']}%")
        if filters.get('contract_id'):
            query += " AND r.contract_id = ?"
            params.append(filters['contract_id'])
    
    query += " ORDER BY r.fecha_inicio DESC, r.id DESC"
    
    cursor.execute(query, params)
    rows = cursor.fetchall()
    conn.close()
    
    data = [["ID", "Número", "Fecha", "Partida", "Llegada", "Transportista"]]
    
    for row in rows:
        data.append([
            str(row['id']),
            row['numero_remision'],
            row['fecha_inicio'],
            row['partida'] or '',
            row['llegada'] or '',
            row['transportista_nombre'] or ''
        ])
    
    table = Table(data)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    
    elements.append(table)
    doc.build(elements)
    buffer.seek(0)
    return buffer.getvalue()


def generate_sales_invoices_excel(filters: Optional[Dict] = None) -> bytes:
    """Genera reporte Excel de facturas de venta"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Facturas de Venta"
    
    headers = ["ID", "Número", "Fecha", "Cliente", "RUC", "Condición", "Total"]
    ws.append(headers)
    
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    
    conn = get_db()
    cursor = conn.cursor()
    
    query = """
        SELECT si.id, si.numero, si.fecha, cl.nombre as cliente_nombre,
               cl.ruc, si.condicion_venta
        FROM sales_invoices si
        LEFT JOIN clients cl ON si.client_id = cl.id
        WHERE 1=1
    """
    params = []
    
    if filters:
        if filters.get('cliente'):
            query += " AND cl.nombre LIKE ?"
            params.append(f"%{filters['cliente']}%")
        if filters.get('contract_id'):
            query += " AND si.contract_id = ?"
            params.append(filters['contract_id'])
    
    query += " ORDER BY si.fecha DESC, si.id DESC"
    
    cursor.execute(query, params)
    rows = cursor.fetchall()
    conn.close()
    
    for row in rows:
        # Calcular total
        conn = get_db()
        cursor = conn.cursor()
        cursor.execute("""
            SELECT SUM(cantidad * precio_unitario) as total
            FROM sales_invoice_items
            WHERE sales_invoice_id = ?
        """, (row['id'],))
        total_row = cursor.fetchone()
        total = total_row['total'] if total_row['total'] else 0
        conn.close()
        
        ws.append([
            row['id'],
            row['numero'],
            row['fecha'],
            row['cliente_nombre'] or '',
            row['ruc'] or '',
            row['condicion_venta'],
            total
        ])
    
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


def generate_sales_invoices_pdf(filters: Optional[Dict] = None) -> bytes:
    """Genera reporte PDF de facturas de venta"""
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    elements = []
    
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=16,
        textColor=colors.HexColor('#2c3e50'),
        spaceAfter=30,
    )
    
    elements.append(Paragraph("Reporte de Facturas de Venta", title_style))
    elements.append(Spacer(1, 0.2*inch))
    
    conn = get_db()
    cursor = conn.cursor()
    
    query = """
        SELECT si.id, si.numero, si.fecha, cl.nombre as cliente_nombre,
               cl.ruc, si.condicion_venta
        FROM sales_invoices si
        LEFT JOIN clients cl ON si.client_id = cl.id
        WHERE 1=1
    """
    params = []
    
    if filters:
        if filters.get('cliente'):
            query += " AND cl.nombre LIKE ?"
            params.append(f"%{filters['cliente']}%")
        if filters.get('contract_id'):
            query += " AND si.contract_id = ?"
            params.append(filters['contract_id'])
    
    query += " ORDER BY si.fecha DESC, si.id DESC"
    
    cursor.execute(query, params)
    rows = cursor.fetchall()
    conn.close()
    
    data = [["ID", "Número", "Fecha", "Cliente", "RUC", "Condición"]]
    
    for row in rows:
        data.append([
            str(row['id']),
            row['numero'],
            row['fecha'],
            row['cliente_nombre'] or '',
            row['ruc'] or '',
            row['condicion_venta']
        ])
    
    table = Table(data)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    
    elements.append(table)
    doc.build(elements)
    buffer.seek(0)
    return buffer.getvalue()

